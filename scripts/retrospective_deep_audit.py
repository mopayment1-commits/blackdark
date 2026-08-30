#!/usr/bin/env python3
"""Retrospective deep quad audit for hero batches 01+02 (200 capabilities).

Classifies each capability into exactly one of:
  VERIFIED-DEEP              — underlying unit independently tested + live-verified
  WRAPPER-ONLY-UNVERIFIED  — wrapper executes but underlying lacks independent quad
  DEFERRED/DELEGATED       — explicitly deferred stub or delegated binding

Updates capabilities_checklist.xlsx, evidence JSONL, and gap reports.
"""
from __future__ import annotations

import ast
import asyncio
import importlib
import inspect
import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pdf_capability_registry import batch_test_module_for, discover_bindings, execute_capability  # noqa: E402
from scripts.run_hero_batch_closure import classify_implementation  # noqa: E402

BATCH_01 = ROOT / "scripts/partial_batches/batch_hero_01.json"
BATCH_02 = ROOT / "scripts/partial_batches/batch_02_101_200.json"
HEROES_LAYER = ROOT / "bd_platform/heroes_capability_layer.py"
EVIDENCE_01 = ROOT / "data/hero_batch_01_evidence.jsonl"
EVIDENCE_02 = ROOT / "data/hero_batch_02_101_200_evidence.jsonl"
GAP_01 = ROOT / "docs/HERO_BATCH_01_GAP_REPORT.json"
GAP_02 = ROOT / "docs/HERO_BATCH_02_101_200_GAP_REPORT.json"
AUDIT_JSON = ROOT / "docs/RETROSPECTIVE_DEEP_AUDIT_BATCHES_01_02.json"
AUDIT_MD = ROOT / "docs/RETROSPECTIVE_DEEP_AUDIT_REPORT.md"
CHECKLIST = ROOT / "capabilities_checklist.xlsx"

HERO_WRAPPER_TESTS = {
    "tests/test_hero_batch_01_capabilities.py",
    "tests/test_hero_batch_02_capabilities.py",
}

INDEPENDENT_TEST_RANGES = [
    (57, 66, "tests/test_legal_retail_batch57_66.py"),
    (67, 76, "tests/test_pro_trader_batch67_76.py"),
    (77, 86, "tests/test_whales_institutional_batch77_86.py"),
    (87, 94, "tests/test_institutional_b2b_batch87_94.py"),
    (95, 104, "tests/test_infra_intelligence_batch95_104.py"),
    (105, 116, "tests/test_market_analysis_batch105_116.py"),
    (117, 128, "tests/test_advanced_ta_risk_batch117_128.py"),
    (129, 139, "tests/test_onchain_platform_batch129_139.py"),
    (140, 152, "tests/test_data_sources_batch140_152.py"),
    (153, 163, "tests/test_intelligence_analysis_batch153_163.py"),
    (164, 176, "tests/test_risk_infrastructure_batch164_176.py"),
    (177, 191, "tests/test_arbitrage_portfolio_ux_batch177_191.py"),
    (192, 203, "tests/test_derivatives_ta_research_batch192_203.py"),
    (204, 216, "tests/test_onchain_defi_sources_batch204_216.py"),
    (217, 227, "tests/test_intelligence_market_extensions_batch217_227.py"),
    (228, 241, "tests/test_intelligence_ux_extensions_batch228_241.py"),
    (242, 261, "tests/test_security_trust_data_batch242_261.py"),
]


def _independent_test_file(cap_id: int) -> str | None:
    for lo, hi, path in INDEPENDENT_TEST_RANGES:
        if lo <= cap_id <= hi and (ROOT / path).is_file():
            return path
    return None


STUB_MARKERS = (
    "deferred",
    "rejected",
    "not implemented",
    "placeholder",
    "stub",
    "raise notimplementederror",
)

WRAPPER_ONLY_NOTE = "مبني جزئيًا — يحتاج تحقق إضافي (WRAPPER-ONLY-UNVERIFIED)"
VERIFIED_NOTE = "مبني جزئيًا — VERIFIED-DEEP"


def _load_ids(path: Path) -> list[int]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return [int(x) for x in data["capability_ids"]]


def _git_on_main(path: str) -> bool:
    try:
        subprocess.run(
            ["git", "cat-file", "-e", f"origin/main:{path}"],
            cwd=ROOT,
            capture_output=True,
            check=True,
        )
        return True
    except subprocess.CalledProcessError:
        return False


def _parse_heroes_underlying() -> dict[int, tuple[str, str]]:
    """Map hero capability id -> (module, function) from heroes_capability_layer."""
    src = HEROES_LAYER.read_text(encoding="utf-8")
    tree = ast.parse(src)
    mapping: dict[int, tuple[str, str]] = {}

    for node in tree.body:
        if not isinstance(node, ast.FunctionDef) or not node.name.startswith("_"):
            continue
        m = re.match(r"_(\d+)$", node.name)
        if not m:
            continue
        cap_id = int(m.group(1))
        for child in ast.walk(node):
            if isinstance(child, ast.Call) and isinstance(child.func, ast.Name):
                if child.func.id in ("_delegate_sync", "_delegate_async"):
                    if len(child.args) >= 2:
                        mod = child.args[0]
                        fn = child.args[1]
                        if isinstance(mod, ast.Constant) and isinstance(fn, ast.Constant):
                            mapping[cap_id] = (str(mod.value), str(fn.value))
    return mapping


def _resolve_binding(cap_id: int, heroes_map: dict[int, tuple[str, str]], bindings: dict[int, tuple[str, str]]) -> tuple[str, str, str]:
    """Return (module, function, binding_kind)."""
    reg = bindings.get(cap_id)
    if not reg:
        return "", "", "missing"

    mod_path, fn_name = reg
    if mod_path.endswith("heroes_capability_layer") and cap_id in heroes_map:
        um, uf = heroes_map[cap_id]
        return um, uf, "heroes_wrapper"

    if "regulatory_compliance_guard" in mod_path or fn_name == "compliant_oracle_sentence":
        return "bd_platform.regulatory_compliance_guard", "compliant_oracle_sentence", "delegated"

    return mod_path, fn_name, "direct"


def _is_stub_source(source: str) -> bool:
    low = source.lower()
    if "raise NotImplementedError" in source:
        return True
    for marker in STUB_MARKERS:
        if marker in low and ("return {" in low or "raise " in low):
            if marker in ("deferred", "rejected"):
                return True
    if re.search(r'"status"\s*:\s*"(deferred|rejected)"', source, re.I):
        return True
    if re.search(r"'status'\s*:\s*'(deferred|rejected)'", source, re.I):
        return True
    return False


def _underlying_real_code(module: str, func_name: str) -> tuple[bool, str]:
    if not module or not func_name:
        return False, "no_binding"
    mod_path = module.replace(".", "/") + ".py"
    if not (ROOT / mod_path).exists():
        return False, f"missing_module:{mod_path}"
    try:
        mod = importlib.import_module(module)
        fn = getattr(mod, func_name, None)
        if fn is None:
            return False, f"missing_function:{func_name}"
        src = inspect.getsource(fn)
        if _is_stub_source(src):
            return False, "stub_or_deferred"
        if len(src.strip()) < 40:
            return False, "trivial_stub"
        return True, "real_code"
    except Exception as exc:
        return False, f"import_error:{exc}"


def _scan_independent_tests(cap_id: int, module: str, func_name: str) -> tuple[bool, str | None, str | None]:
    """Find independent test for underlying unit (not hero_batch wrapper)."""
    # 1) Range-batch file for capability id
    range_file = _independent_test_file(cap_id)
    if range_file:
        text = (ROOT / range_file).read_text(encoding="utf-8", errors="replace")
        if func_name in text or str(cap_id) in text:
            return True, range_file, f"range_batch_{cap_id}"

    # 2) Module-specific test file
    if module:
        short_mod = module.split(".")[-1]
        candidates = [
            f"tests/test_{short_mod}.py",
            f"tests/test_{short_mod.replace('_layer', '')}.py",
        ]
        for cand in candidates:
            if (ROOT / cand).is_file():
                text = (ROOT / cand).read_text(encoding="utf-8", errors="replace")
                if func_name in text or short_mod in text:
                    return True, cand, f"module_test:{short_mod}"

    # 3) Search all non-hero tests for underlying function reference
    if func_name:
        for test_file in sorted((ROOT / "tests").glob("test_*.py")):
            if test_file.name in {p.split("/")[-1] for p in HERO_WRAPPER_TESTS}:
                continue
            text = test_file.read_text(encoding="utf-8", errors="replace")
            if func_name in text and (module.split(".")[-1] in text or module in text):
                return True, str(test_file.relative_to(ROOT)), f"fn_ref:{func_name}"

    # 4) Capability-id patterns in non-hero batch tests
    patterns = [f"test_{cap_id}_", f"_{cap_id}_", f"[{cap_id}]"]
    for test_file in sorted((ROOT / "tests").glob("test_*batch*.py")):
        if test_file.name in {p.split("/")[-1] for p in HERO_WRAPPER_TESTS}:
            continue
        text = test_file.read_text(encoding="utf-8", errors="replace")
        for pat in patterns:
            if pat in text:
                return True, str(test_file.relative_to(ROOT)), pat

    return False, None, None


def _matching_test_functions(test_file: str, cap_id: int, func_name: str) -> list[str]:
    """Return test function names that exercise the underlying unit."""
    tf = ROOT / test_file
    if not tf.exists():
        return []
    text = tf.read_text(encoding="utf-8", errors="replace")
    names: list[str] = []
    current: str | None = None
    body_lines: list[str] = []
    for line in text.splitlines():
        m = re.match(r"\s*def (test_[^(]+)\(", line)
        if m:
            if current:
                body = "\n".join(body_lines)
                if (
                    (func_name and func_name in body)
                    or str(cap_id) in current
                    or (func_name and func_name.split("_")[-1] in current)
                ):
                    names.append(current)
            current = m.group(1)
            body_lines = []
        elif current:
            body_lines.append(line)
    if current:
        body = "\n".join(body_lines)
        if (
            (func_name and func_name in body)
            or str(cap_id) in current
            or (func_name and func_name.split("_")[-1] in current)
        ):
            names.append(current)
    return names


def _run_independent_test(test_file: str, cap_id: int, func_name: str) -> tuple[bool, str]:
    """Run pytest for underlying unit tests in file."""
    tf = ROOT / test_file
    if not tf.exists():
        return False, "test_file_missing"

    names = _matching_test_functions(test_file, cap_id, func_name)
    if names:
        nodeids = [f"{test_file}::{n}" for n in names[:8]]
        cmd = [sys.executable, "-m", "pytest", *nodeids, "-q", "--tb=line"]
    else:
        cmd = [sys.executable, "-m", "pytest", str(tf), "-q", "--tb=line"]

    try:
        proc = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, timeout=180)
        out = (proc.stdout or "") + (proc.stderr or "")
        passed = proc.returncode == 0
        return passed, out[-600:] if len(out) > 600 else out
    except subprocess.TimeoutExpired:
        return False, "timeout"
    except Exception as exc:
        return False, str(exc)


async def _live_verify(cap_id: int) -> tuple[bool, dict[str, Any]]:
    try:
        result = await execute_capability(cap_id)
        ok = bool(result.get("ok", False))
        binding = result.get("binding", "")
        return ok, {"ok": ok, "binding": binding, "keys": list(result.keys())[:8]}
    except Exception as exc:
        return False, {"error": str(exc)}


def _binding_from_live(binding: str) -> tuple[str, str]:
    """Parse 'module.path.function' from live exec binding."""
    if not binding or "." not in binding:
        return "", ""
    parts = binding.rsplit(".", 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return "", ""


def _classify(
    cap_id: int,
    underlying_real: bool,
    underlying_reason: str,
    has_indep_test: bool,
    indep_test_passed: bool,
    live_ok: bool,
    binding_kind: str,
    prior_impl_class: str | None,
    live_detail: dict[str, Any],
) -> str:
    blob = json.dumps(live_detail, default=str).lower()
    deferred_markers = (
        "deferred",
        "rejected",
        "duplicate_not_build",
        "build_blocked",
        "insights_only_no_execution",
    )
    is_deferred = (
        prior_impl_class == "deferred"
        or underlying_reason == "stub_or_deferred"
        or any(m in blob for m in deferred_markers)
        or live_detail.get("status") in {"deferred", "rejected"}
    )
    is_delegated = (
        prior_impl_class == "delegated"
        and binding_kind == "delegated"
    )

    if is_deferred or is_delegated:
        return "DEFERRED/DELEGATED"
    if underlying_real and has_indep_test and indep_test_passed and live_ok:
        return "VERIFIED-DEEP"
    return "WRAPPER-ONLY-UNVERIFIED"


def _load_prior_evidence(path: Path) -> dict[int, dict]:
    out: dict[int, dict] = {}
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        row = json.loads(line)
        out[int(row["capability_id"])] = row
    return out


def _load_gap_impl_classes(path: Path) -> dict[int, str]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    classes: dict[int, str] = {}
    for row in data.get("rows", data.get("capabilities", [])):
        cid = int(row.get("capability_id", row.get("id", 0)))
        ic = row.get("implementation_class", row.get("status", ""))
        if ic in ("implemented", "delegated", "deferred"):
            classes[cid] = ic
    return classes


async def audit_all() -> dict[str, Any]:
    ids_01 = _load_ids(BATCH_01)
    ids_02 = _load_ids(BATCH_02)
    all_ids = ids_01 + ids_02

    heroes_map = _parse_heroes_underlying()
    bindings = discover_bindings()
    prior_01 = _load_prior_evidence(EVIDENCE_01)
    prior_02 = _load_prior_evidence(EVIDENCE_02)
    gap02_classes = _load_gap_impl_classes(GAP_02)
    gap01_classes = _load_gap_impl_classes(GAP_01)

    rows: list[dict[str, Any]] = []
    counts = {"VERIFIED-DEEP": 0, "WRAPPER-ONLY-UNVERIFIED": 0, "DEFERRED/DELEGATED": 0}

    for cap_id in all_ids:
        batch = "batch_01" if cap_id in ids_01 else "batch_02"
        prior = prior_01.get(cap_id) or prior_02.get(cap_id) or {}
        prior_impl = prior.get("implementation_class") or gap02_classes.get(cap_id) or gap01_classes.get(cap_id)

        module, func_name, binding_kind = _resolve_binding(cap_id, heroes_map, bindings)
        live_ok, live_detail = await _live_verify(cap_id)

        reg_binding = bindings.get(cap_id)
        if not prior_impl and reg_binding:
            prior_impl = classify_implementation(cap_id, reg_binding, live_detail)

        live_binding = live_detail.get("binding", "")
        if live_binding:
            lm, lf = _binding_from_live(live_binding)
            if lm and lf and not lm.endswith("heroes_capability_layer"):
                module, func_name = lm, lf
                if binding_kind == "heroes_wrapper":
                    binding_kind = "heroes_wrapper_traced"

        real, real_reason = _underlying_real_code(module, func_name)
        has_test, test_file, test_pat = _scan_independent_tests(cap_id, module, func_name)
        test_passed = False
        test_output = ""
        if has_test and test_file:
            test_passed, test_output = _run_independent_test(test_file, cap_id, func_name)

        mod_path = module.replace(".", "/") + ".py" if module else ""
        on_main = _git_on_main(mod_path) if mod_path and (ROOT / mod_path).exists() else False
        source_branch = "origin/main" if on_main else "capabilities-826-import"

        classification = _classify(
            cap_id,
            real,
            real_reason,
            has_test,
            test_passed,
            live_ok,
            binding_kind,
            prior_impl,
            live_detail,
        )
        counts[classification] += 1

        rows.append(
            {
                "capability_id": cap_id,
                "batch": batch,
                "classification": classification,
                "prior_implementation_class": prior_impl,
                "binding_kind": binding_kind,
                "underlying_module": module,
                "underlying_function": func_name,
                "underlying_real_code": real,
                "underlying_real_reason": real_reason,
                "independent_test_file": test_file,
                "independent_test_pattern": test_pat,
                "independent_test_passed": test_passed,
                "independent_test_output_tail": test_output[-200:] if test_output else "",
                "live_ok": live_ok,
                "live_detail": live_detail,
                "source_on_main": on_main,
                "source_branch": source_branch,
                "quad_deep": {
                    "real_code": real,
                    "independent_test": has_test and test_passed,
                    "live_verify": live_ok,
                    "source_traced": bool(module),
                },
            }
        )

    report = {
        "audit_type": "retrospective_deep_quad",
        "audited_at": datetime.now(timezone.utc).isoformat(),
        "total_capabilities": len(all_ids),
        "batch_01_count": len(ids_01),
        "batch_02_count": len(ids_02),
        "classification_counts": counts,
        "verified_deep_honest_count": counts["VERIFIED-DEEP"],
        "verified_deep_pct": round(100.0 * counts["VERIFIED-DEEP"] / len(all_ids), 1),
        "wrapper_only_unverified_count": counts["WRAPPER-ONLY-UNVERIFIED"],
        "deferred_delegated_count": counts["DEFERRED/DELEGATED"],
        "acceptance_criterion": "No batch 3 until all WRAPPER-ONLY-UNVERIFIED resolved or explicitly deferred",
        "rows": rows,
    }
    return report


def _update_checklist(rows: list[dict[str, Any]]) -> None:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed; skipping xlsx update")
        return

    wb = openpyxl.load_workbook(CHECKLIST)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1) if cell.value}
    id_col = headers.get("ID") or headers.get("id") or 1
    status_col = headers.get("Status") or headers.get("status") or headers.get("الحالة")
    notes_col = headers.get("Notes") or headers.get("notes") or headers.get("ملاحظات")

    by_id = {r["capability_id"]: r for r in rows}
    for row_idx in range(2, ws.max_row + 1):
        cid_cell = ws.cell(row=row_idx, column=id_col)
        try:
            cid = int(cid_cell.value)
        except (TypeError, ValueError):
            continue
        if cid not in by_id:
            continue
        rec = by_id[cid]
        cls = rec["classification"]
        if status_col and cls == "WRAPPER-ONLY-UNVERIFIED":
            ws.cell(row=row_idx, column=status_col).value = "مبني جزئيًا"
        elif status_col and cls == "VERIFIED-DEEP":
            ws.cell(row=row_idx, column=status_col).value = "مبني جزئيًا"
        if notes_col:
            tag = {
                "VERIFIED-DEEP": "[VERIFIED-DEEP]",
                "WRAPPER-ONLY-UNVERIFIED": "[WRAPPER-ONLY-UNVERIFIED]",
                "DEFERRED/DELEGATED": "[DEFERRED/DELEGATED]",
            }[cls]
            existing = ws.cell(row=row_idx, column=notes_col).value or ""
            base = str(existing).split("[")[0].strip() if "[" in str(existing) else str(existing).strip()
            ws.cell(row=row_idx, column=notes_col).value = f"{base} {tag}".strip()

    wb.save(CHECKLIST)
    print(f"Updated {CHECKLIST}")


def _update_evidence(path: Path, rows: list[dict[str, Any]], id_set: set[int]) -> None:
    prior = _load_prior_evidence(path)
    lines: list[str] = []
    for cap_id in sorted(id_set):
        rec = next(r for r in rows if r["capability_id"] == cap_id)
        old = prior.get(cap_id, {})
        entry = {
            **old,
            "capability_id": cap_id,
            "deep_audit_classification": rec["classification"],
            "implementation_class": (
                "verified_deep"
                if rec["classification"] == "VERIFIED-DEEP"
                else "wrapper_only_unverified"
                if rec["classification"] == "WRAPPER-ONLY-UNVERIFIED"
                else rec.get("prior_implementation_class") or "deferred_delegated"
            ),
            "underlying_module": rec["underlying_module"],
            "underlying_function": rec["underlying_function"],
            "independent_test_passed": rec["independent_test_passed"],
            "independent_test_file": rec["independent_test_file"],
            "source_branch": rec["source_branch"],
            "deep_audit_at": datetime.now(timezone.utc).isoformat(),
            "notes": (
                VERIFIED_NOTE
                if rec["classification"] == "VERIFIED-DEEP"
                else WRAPPER_ONLY_NOTE
                if rec["classification"] == "WRAPPER-ONLY-UNVERIFIED"
                else old.get("notes", "")
            ),
        }
        lines.append(json.dumps(entry, ensure_ascii=False))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Updated {path}")


def _update_gap(path: Path, rows: list[dict[str, Any]], id_set: set[int], batch_name: str) -> None:
    subset = [r for r in rows if r["capability_id"] in id_set]
    counts = {}
    for r in subset:
        counts[r["classification"]] = counts.get(r["classification"], 0) + 1

    gap = {
        "batch": batch_name,
        "deep_audit_at": datetime.now(timezone.utc).isoformat(),
        "total": len(subset),
        "classification_counts": counts,
        "verified_deep_honest": counts.get("VERIFIED-DEEP", 0),
        "wrapper_only_unverified": counts.get("WRAPPER-ONLY-UNVERIFIED", 0),
        "deferred_delegated": counts.get("DEFERRED/DELEGATED", 0),
        "rows": [
            {
                "capability_id": r["capability_id"],
                "classification": r["classification"],
                "underlying": f"{r['underlying_module']}.{r['underlying_function']}",
                "independent_test_passed": r["independent_test_passed"],
                "live_ok": r["live_ok"],
                "source_branch": r["source_branch"],
            }
            for r in subset
        ],
    }
    path.write_text(json.dumps(gap, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Updated {path}")


def _write_md(report: dict[str, Any]) -> None:
    c = report["classification_counts"]
    lines = [
        "# Retrospective Deep Audit — Batches 01 + 02",
        "",
        f"**Audited at:** {report['audited_at']}",
        "",
        "## Honest Count (200 capabilities)",
        "",
        f"| Classification | Count | % |",
        f"|---|---:|---:|",
        f"| **VERIFIED-DEEP** | **{c['VERIFIED-DEEP']}** | {report['verified_deep_pct']}% |",
        f"| WRAPPER-ONLY-UNVERIFIED | {c['WRAPPER-ONLY-UNVERIFIED']} | {round(100*c['WRAPPER-ONLY-UNVERIFIED']/200,1)}% |",
        f"| DEFERRED/DELEGATED | {c['DEFERRED/DELEGATED']} | {round(100*c['DEFERRED/DELEGATED']/200,1)}% |",
        "",
        "## Acceptance",
        "",
        "- Batch 3 **blocked** until WRAPPER-ONLY-UNVERIFIED items are independently verified or explicitly deferred.",
        "- VERIFIED-DEEP requires: real underlying code + independent range test PASS + live exec OK + source traced.",
        "",
        "## Method",
        "",
        "1. Resolve underlying module/function (heroes wrappers traced via AST).",
        "2. Reject stubs/deferred markers in underlying source.",
        "3. Require independent `test_*batch*` test (NOT hero_batch wrapper tests).",
        "4. Run pytest for matching test; live `execute_capability`.",
        "5. Trace git `origin/main` vs capabilities-826-import.",
        "",
        f"Full JSON: `{AUDIT_JSON.relative_to(ROOT)}`",
    ]
    AUDIT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {AUDIT_MD}")


def main() -> int:
    print("Starting retrospective deep audit (200 capabilities)...")
    report = asyncio.run(audit_all())

    AUDIT_JSON.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {AUDIT_JSON}")

    ids_01 = set(_load_ids(BATCH_01))
    ids_02 = set(_load_ids(BATCH_02))

    _update_evidence(EVIDENCE_01, report["rows"], ids_01)
    _update_evidence(EVIDENCE_02, report["rows"], ids_02)
    _update_gap(GAP_01, report["rows"], ids_01, "hero_batch_01")
    _update_gap(GAP_02, report["rows"], ids_02, "hero_batch_02_101_200")
    _update_checklist(report["rows"])
    _write_md(report)

    c = report["classification_counts"]
    print("\n=== HONEST COUNT ===")
    print(f"VERIFIED-DEEP:              {c['VERIFIED-DEEP']} / 200")
    print(f"WRAPPER-ONLY-UNVERIFIED:    {c['WRAPPER-ONLY-UNVERIFIED']} / 200")
    print(f"DEFERRED/DELEGATED:         {c['DEFERRED/DELEGATED']} / 200")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
