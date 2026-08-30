#!/usr/bin/env python3
"""Audit capabilities from attached PDF checklist only — no CAP978."""

from __future__ import annotations

import json
import re
import subprocess
from pathlib import Path
from typing import Any

import pymupdf as fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
PDF = Path("/home/ubuntu/.cursor/projects/workspace/uploads/capabilities_checklist_66b7.pdf")
OUT = ROOT / "capabilities_checklist.xlsx"

# ── curated evidence (code paths only, no catalog) ─────────────────────────
EXACT: dict[int, tuple[str, str]] = {
    1: ("مبني وشغال فعليًا", "bd_platform/footprint_analytics.py:footprint_snapshot + market_analysis_layer compute_imbalance_delta_113"),
    2: ("مبني جزئيًا", "trade_simulator.py + grid_bot.py (paper/sim only; لا محرك paper-trading broker حقيقي)"),
    3: ("مبني وشغال فعليًا", "dashboard.py /api/journal + pro_trader_layer add_journal_entry_76"),
    4: ("مبني وشغال فعليًا", "aggregator.py + live_book_hub.py + market_analysis_layer"),
    5: ("مبني وشغال فعليًا", "cap646/handlers/market.py depth_level L1 (id 508) + live_book_hub"),
    6: ("مبني وشغال فعليًا", "cap646/handlers/market.py depth_level L2 (id 509)"),
    7: ("مبني وشغال فعليًا", "cap646/handlers/market.py depth_level L3 (id 510)"),
    8: ("مبني جزئيًا", "trade_simulator.py + grid_bot.py + strategy_simulator_195 (simulation فقط)"),
    12: ("مبني وشغال فعليًا", "instant_alert_engine.py + in_app_alerts.py"),
    13: ("مبني وشغال فعليًا", "instant_alert_engine.py + cap646/handlers/alerts.py"),
    14: ("مبني وشغال فعليًا", "bd_platform/whale_story.py + pro_trader_layer build_whale_narrative_71"),
    15: ("مبني جزئيًا", "onchain_platform_layer scan_flash_loan_vulnerabilities_132 (scan فقط)"),
    21: ("مبني وشغال فعليًا", "in_app_alerts.py inbox + dashboard alert routes"),
    36: ("مبني وشغال فعليًا", "execution_engine.py trigger_panic + risk_manager"),
    37: ("مبني جزئيًا", "execution_engine.py موجود لكن AUTO_EXECUTION_ENABLED=false وDRY_RUN=true افتراضيًا"),
    41: ("مبني جزئيًا", "execution_engine.py + arbitrage_engine (تنفيذ تلقائي معطّل افتراضيًا؛ execution_rejected_layer"),
    43: ("مبني وشغال فعليًا", "bd_platform/mvrv_realignment.py compute_mvrv_realignment"),
    44: ("مبني وشغال فعليًا", "bd_platform/alpha_engine.py compute_alpha_signal"),
    45: ("مبني جزئيًا", "blackdark/ingestion/arkham_connector.py (proxy/input؛ يعتمد مفتاح خارجي)"),
    47: ("مبني جزئيًا", "whales_institutional_layer build_exchange_health_80 (تحليلي؛ ليس شهادة مستقلة)"),
    49: ("مبني جزئيًا", "risk_manager + drawdown_guard (حماية؛ ليس flash-crash engine مخصص بالكامل)"),
    50: ("مبني جزئيًا", "bd_platform/onchain_advanced.py MVRV proxies + lookintobitcoin_macro"),
    217: ("مبني جزئيًا", "buyer_model_card.py + docs/AI_FINANCIAL_MODEL_DESIGN.md (توثيق؛ ليس MRM رسمي)"),
    55: ("مبني وشغال فعليًا", "due_diligence_bundle.py build_full_due_diligence_bundle + due_diligence.py"),
    69: ("مبني وشغال فعليًا", "net_edge_truth.py compute_net_edge_truth + signal_registry"),
    17: ("مبني وشغال فعليًا", "instant_alert_engine.py + in_app_alerts.py + telegram_agent.py"),
    113: ("غير موجود إطلاقًا", "لا وحدة M&A intelligence في الكود"),
    221: ("مبني جزئيًا", "committee_one_pager.py + acquirer_evidence_pack.py (تصدير جزئي؛ ليس تقارير لجنة كاملة)"),
    380: ("غير موجود إطلاقًا", "لا قائمة عملات إيداع مفتوحة/مغلقة في الكود"),
    381: ("غير موجود إطلاقًا", "لا قائمة عملات سحب مفتوحة/مغلقة في الكود"),
    382: ("مبني جزئيًا", "ai_oracle.py evaluate_opportunity + dashboard /oracle/{symbol} (sanitized publicly)"),
    691: ("مبني جزئيًا", "data_sources_registry.py + exchange connectors (KuCoin جزئي عبر registry)"),
    745: ("مبني جزئيًا", "subscription_analytics.py + billing analytics (ليس visitor counter كامل)"),
    754: ("مبني جزئيًا", "docs/ops/INCIDENT_RESPONSE.md + data/institutional_assurance/incident_response.json"),
    409: ("مبني جزئيًا", "news_classifier + intelligence layers (feed؛ ليس منتج QuickTake كامل)"),
    702: ("مبني جزئيًا", "mcp server references + graphql_schema (جزئي)"),
    703: ("مبني وشغال فعليًا", "platform_api.py + public_api_docs.py REST surfaces"),
    706: ("مبني جزئيًا", "dbt_connector.py (يتطلب إعداد dbt خارجي)"),
    809: ("مبني جزئيًا", "dashboard.py I DON'T KNOW dead-zone + constitution_gates (ليس محرك epistemic مستقل)"),
    810: ("مبني جزئيًا", "enterprise_sso.py + org_rbac.py (SSO/RBAC جزئي؛ ليس IdP enterprise كامل)"),
    811: ("مبني جزئيًا", "gdpr_service.py + legal_commercial_layer gdpr_compliance_status_58"),
    812: ("مبني وشغال فعليًا", "bd_platform/trulens_eval.py + decision_certificate.py"),
    813: ("مبني جزئيًا", "constitution_gates.py + dimension_conflict_guard (شروط إبطال جزئية)"),
    814: ("مبني وشغال فعليًا", "oracle_track_record.py + public_accuracy_ledger"),
    815: ("مبني وشغال فعليًا", "oracle_audit_chain.py + locked_predictions.py"),
    816: ("مبني جزئيًا", "dimension_conflict_guard abstain/veto + ml/drift_monitor OOD fail-closed"),
}

# (regex on name lower, status, evidence) — checked before generic search
PATTERN_RULES: list[tuple[str, str, str]] = [
    (r"telegram|whatsapp|discord", "مبني وشغال فعليًا", "bd_platform/telegram_agent.py + intelligence_analysis_layer #161"),
    (r"watchlist", "مبني جزئيًا", "cap646/ui_pages.py watchlists + security_trust_data_layer list_etherscan_watchlist_246"),
    (r"order book|market depth|orderbook", "مبني وشغال فعليًا", "aggregator.py + live_book_hub.py + cap646/handlers/market.py"),
    (r"orderflow|order flow|imbalance", "مبني وشغال فعليًا", "bd_platform/footprint_analytics.py + market_analysis_layer"),
    (r"alert orchestration|unified alert", "مبني جزئيًا", "alert_service.py + instant_alert_engine.py (لا orchestrator مستقل)"),
    (r"custom metric alert|custom alert", "مبني جزئيًا", "bd_platform/ifttt_rules.py + instant_alert_engine.py"),
    (r"logging|metrics|tracing", "مبني جزئيًا", "institutional_assurance.py + uptime_probe_loop + centralized logs"),
    (r"network growth", "مبني وشغال فعليًا", "bd_platform/footprint_analytics.py footprint_snapshot"),
    (r"single-sentence|financial oracle|oracle", "مبني جزئيًا", "ai_oracle.py + oracle_unified.py (داخلي Buy Now؛ عام sanitized عبر regulatory_compliance_guard)"),
    (r"arbitrag", "مبني جزئيًا", "arbitrage_engine.py + intelligence_analysis_layer #153"),
    (r"flash loan", "مبني جزئيًا", "onchain_platform_layer scan_flash_loan_vulnerabilities_132"),
    (r"whale", "مبني وشغال فعليًا", "bd_platform/whale_story.py + whale_tracker.py"),
    (r"stablecoin", "مبني جزئيًا", "bd_platform/onchain_hub.py + market_context stablecoin handling"),
    (r"tvl", "مبني وشغال فعليًا", "bd_platform/onchain_hub.py defillama + free_tier"),
    (r"unlock", "مبني وشغال فعليًا", "bd_platform/token_unlocks.py unlock_calendar"),
    (r"gas cost|gas ", "مبني وشغال فعليًا", "gas_oracle.py + cap646/fallbacks resolve_gas_usd"),
    (r"funding rate|open interest|liquidation", "مبني وشغال فعليًا", "bd_platform/derivatives_hub.py + liquidation_radar.py"),
    (r"cvd|taker|maker", "مبني وشغال فعليًا", "bd_platform/derivatives_ta_research_layer.py"),
    (r"portfolio", "مبني وشغال فعليًا", "bd_platform/portfolio_rebalancer.py + database portfolio tables"),
    (r"backtest", "مبني جزئيًا", "ml/market_replay_bootstrap.py + derivatives_ta strategy_simulator_195"),
    (r"chart|tradingview", "مبني وشغال فعليًا", "bd_platform/tradingview_bridge.py chart_config"),
    (r"sentiment|fear.?greed", "مبني وشغال فعليًا", "sentiment_engine.py + sentiment_gate.py"),
    (r"provenance|lineage|data quality", "مبني جزئيًا", "data_provenance_score.py + blackdark/data/provenance.py (لا visualization)"),
    (r"sso|rbac|mfa|2fa", "مبني جزئيًا", "enterprise_sso.py + org_rbac.py + admin_mfa.py"),
    (r"gdpr|data residency|right to be forgotten", "مبني جزئيًا", "gdpr_service.py + legal_commercial_layer #58"),
    (r"kyc|aml|sanction", "مبني جزئيًا", "institutional_commerce.py + legal_commercial_layer evaluate_aml_gate_59"),
    (r"model card|model documentation|منهجية النموذج", "مبني جزئيًا", "buyer_model_card.py + /model-card (ليس MRM رسمي)"),
    (r"drift|ood|concept", "مبني جزئيًا", "ml/drift_monitor.py (لا حوكمة MRM مستقلة)"),
    (r"pentest|security verification", "مبني جزئيًا", "pentest_attestation.py (قالب؛ verify_pentest_attestation=False)"),
    (r"incident response|استجابة حوادث", "مبني جزئيًا", "docs/ops/INCIDENT_RESPONSE.md + institutional_assurance incident_response.json"),
    (r"reconcil|تسوية", "مبني جزئيًا", "market_context.py multi-source failover (ليس reconciliation engine)"),
    (r"exchange health|certification", "مبني جزئيًا", "whales_institutional_layer build_exchange_health_80"),
    (r"netflow|exchange flow", "مبني جزئيًا", "bd_platform/onchain_hub.py + exchange outflow intelligence layers"),
    (r"slippage", "مبني وشغال فعليًا", "bd_platform/slippage_tolerance_optimizer.py + slippage_guard.py"),
    (r"grid bot|dca", "مبني وشغال فعليًا", "bd_platform/grid_bot.py"),
    (r"pairs trading|stat.?arb|statistical", "مبني وشغال فعليًا", "bd_platform/pairs_trading.py + intelligence_analysis_layer"),
    (r"options", "مبني جزئيًا", "options_fetcher.py + paper_options_oms.py"),
    (r"perp|futures|derivatives", "مبني جزئيًا", "bd_platform/derivatives_hub.py + perp_dex_fetcher"),
    (r"defi|dex |uniswap|aave", "مبني جزئيًا", "defi_arbitrage_engine.py + bd_platform/onchain_hub.py"),
    (r"on-?chain|onchain|wallet profiler", "مبني جزئيًا", "bd_platform/onchain_hub.py + onchain_tracker.py + free_tier_capabilities"),
    (r"etf", "مبني جزئيًا", "bd_platform/free_tier_capabilities etf_flow + etf intelligence modules"),
    (r"macro|fred|dxy", "مبني جزئيًا", "bd_platform/onchain_hub lookintobitcoin_macro + macro layers"),
    (r"news|event monitor|quicktake", "مبني جزئيًا", "bd_platform/news_classifier.py + market_event_library.py"),
    (r"api |rest api|graphql|mcp", "مبني جزئيًا", "platform_api.py + graphql_schema.py + mcp references"),
    (r"bigquery|snowflake|datashare|dbt|warehouse", "مبني جزئيًا", "bigquery_export.py + dbt_connector.py + data_lake.py"),
    (r"binance|okx|bybit|100 منصة|105 عملة", "مبني جزئيًا", "platform_universe.py + universe_rollout.py (تغطية جزئية؛ ليس 100 منصة مؤكدة)"),
    (r"kill rate|anti.?hype|proof", "مبني وشغال فعليًا", "kill_rate_board.py + anti_hype_mode.py + proof_arena.py"),
    (r"accuracy ledger|track record|سجل دقة", "مبني وشغال فعليًا", "oracle_track_record.py + public_accuracy_ledger"),
    (r"timestamp|hashed|إثبات زمني", "مبني وشغال فعليًا", "oracle_audit_chain.py + locked_predictions.py"),
    (r"epistemic|i don.?t know|الامتناع", "مبني جزئيًا", "dashboard.py I DON'T KNOW + dimension_conflict_guard"),
    (r"institutional|b2b|white.?label", "مبني جزئيًا", "bd_platform/institutional_b2b_layer.py + b2b_websocket_hub.py"),
    (r"billing|subscription|stripe", "مبني جزئيًا", "billing_service.py + billing/subscription_engine.py"),
    (r"audit trail|audit log", "مبني وشغال فعليًا", "audit_registry.py + security_trust_data_layer audit_log_id_242"),
    (r"encryption|vault|hsm|secrets", "مبني جزئيًا", "secrets_vault.py Fernet + bd_platform/vault_client.py (ليس HSM)"),
    (r"load test|capacity|chaos", "مبني جزئيًا", "scripts/load_test*.py + scale_readiness.py (آخر run 2026-08-12)"),
    (r"wcag|accessibility|screen reader", "غير موجود إطلاقًا", "بعض aria-label فقط؛ لا اختبار WCAG آلي"),
    (r"feature flag|canary", "مبني جزئيًا", "config.py env toggles + universe_rollout.py (ليس canary deployment)"),
    (r"due diligence|عناية واجبة", "مبني وشغال فعليًا", "due_diligence_bundle.py + due_diligence.py"),
    (r"net.?edge|net edge truth", "مبني وشغال فعليًا", "net_edge_truth.py compute_net_edge_truth"),
    (r"sharpe|sortino|calmar", "مبني جزئيًا", "bd_platform/derivatives_ta_research_layer.py strategy metrics"),
    (r"mindshare|correlation.*mind", "مبني جزئيًا", "bd_platform/correlation_mindshare.py"),
    (r"long.?short|ls ratio", "مبني جزئيًا", "bd_platform/derivatives_hub.py long_short metrics"),
    (r"dev health|developer health", "مبني جزئيًا", "dev_health_score.py"),
    (r"kucoin|binance|okx|bybit|coinbase|kraken|bitfinex|gate\.io|mexc", "مبني جزئيًا", "data_sources_registry.py + exchange connectors (تغطية جزئية)"),
    (r"visitor|مشتركين|اشتراك المستخدم", "مبني جزئيًا", "subscription_analytics.py + billing_service.py"),
    (r"إيداع|سحب|deposit|withdrawal", "مبني جزئيًا", "arbitrage_portfolio_ux_layer withdrawal_suspension_alert_191 (تنبيه؛ ليس custodial)"),
    (r"تقرير.*لجنة|investment committee", "مبني جزئيًا", "committee_one_pager.py + acquirer_evidence_pack.py"),
]

STRONG_FILE_EVIDENCE: dict[str, tuple[str, str]] = {
    "net_edge_truth.py": ("مبني وشغال فعليًا", "net_edge_truth.compute_net_edge_truth"),
    "due_diligence_bundle.py": ("مبني وشغال فعليًا", "due_diligence_bundle.build_full_due_diligence_bundle"),
    "due_diligence.py": ("مبني جزئيًا", "due_diligence.due_diligence_report"),
    "dev_health_score.py": ("مبني جزئيًا", "dev_health_score module"),
    "data_sources_registry.py": ("مبني جزئيًا", "data_sources_registry.py (تسجيل مصادر؛ ليس SLA monitoring كامل)"),
    "cap646/handlers/institutional.py": ("مبني جزئيًا", "institutional handler surfaces"),
    "sentiment_manipulation_guard.py": ("مبني وشغال فعليًا", "sentiment_manipulation_guard spoof detection"),
    "billing/subscription_engine.py": ("مبني جزئيًا", "billing/subscription_engine.py"),
    "coverage_honesty.py": ("مبني وشغال فعليًا", "coverage_honesty.build_coverage_honesty_board"),
    "market_rankings.py": ("مبني وشغال فعليًا", "bd_platform/market_rankings.py"),
    "acquirer_evidence_pack.py": ("مبني جزئيًا", "acquirer_evidence_pack.py"),
}

IGNORE_HIT_SUBSTR = ("scripts/audit_pdf_capabilities_checklist.py",)

REJECTED_KW = (
    "تنفيذ تلقائي",
    "تنفيذ األربيتراج",
    "إرسال تنبيهات أو تنفيذ",
    "safe liquidation",
    "smart execution",
    "brokerage",
    "panic-button safe",
)

PARTIAL_KW = (
    "paper trading",
    "simulation",
    "simulator",
    "proxy",
    "free platform",
    "rejected",
    "insight only",
    "stub",
)

WORKING_FILES = (
    "tests/test_",
    "dashboard.py",
    "platform_api.py",
    "bd_platform/",
    "execution_engine.py",
    "arbitrage_engine.py",
    "instant_alert_engine.py",
    "ml/",
)


def parse_pdf() -> list[dict[str, Any]]:
    doc = fitz.open(str(PDF))
    full = "\n".join(page.get_text() for page in doc)
    full = full.replace("يظهر للمستخدم", "\nيظهر للمستخدم\n").replace("قدرة داخلية", "\nقدرة داخلية\n")
    lines = [ln.strip() for ln in full.splitlines() if ln.strip() and ln.strip() not in {"#اسم الميزة/القدرة", "النوع"}]
    entries: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for ln in lines:
        if ln in ("يظهر للمستخدم", "قدرة داخلية"):
            if current:
                current["type"] = ln
                text = re.sub(r"\s+", " ", current["text"]).strip()
                m = re.match(r"^(\d{1,4})\s*(.+)$", text) or re.match(r"^(\d{1,4})(.+)$", text)
                if m:
                    current["id"] = int(m.group(1))
                    current["name"] = m.group(2).strip()
                entries.append(current)
                current = None
            continue
        current = {"text": ln} if current is None else current.update(text=current["text"] + " " + ln) or current
    return entries


def _tokens(name: str) -> list[str]:
    name = re.sub(r"[^\w\s\u0600-\u06FF]", " ", name)
    parts = [p for p in name.split() if len(p) >= 3]
  # drop ultra-common
    stop = {"the", "and", "for", "with", "free", "data", "engine", "intelligence", "platform", "api", "قدرة", "داخلية", "يظهر", "للمستخدم"}
    return [p for p in parts if p.lower() not in stop][:6]


def rg_search(pattern: str, limit: int = 8) -> list[str]:
    try:
        proc = subprocess.run(
            ["rg", "-l", "-i", pattern, "--glob", "*.py", str(ROOT)],
            capture_output=True,
            text=True,
            timeout=8,
        )
        lines = [ln.strip() for ln in proc.stdout.splitlines() if ln.strip()][:limit]
        return lines
    except Exception:
        return []


def classify(entry: dict[str, Any]) -> str:
    cid = entry["id"]
    name = entry["name"]
    nl = name.lower()

    if cid in EXACT:
        status, evidence = EXACT[cid]
        return f"{status} — {evidence}"

    for pat, status, evidence in PATTERN_RULES:
        if re.search(pat, nl, re.I) or re.search(pat, name, re.I):
            return f"{status} — {evidence}"

    for path, (status, evidence) in STRONG_FILE_EVIDENCE.items():
        if path in nl or any(tok.lower() in path.lower() for tok in _tokens(name)):
            p = ROOT / path
            if p.exists():
                return f"{status} — {evidence}"

    if any(k in name or k in nl for k in REJECTED_KW):
        hits = rg_search(re.escape(name[:40]) if len(name) > 8 else name.split()[0])
        ev = hits[0] if hits else "execution_rejected_layer.py"
        return f"مبني جزئيًا — مرفوض/معطّل تنفيذيًا أو insight-only؛ مرجع: {ev}"

    tokens = _tokens(name)
    hits: list[str] = []
    for tok in tokens:
        hits.extend(rg_search(re.escape(tok), limit=4))
    hits = [h for h in dict.fromkeys(hits) if not any(ig in h for ig in IGNORE_HIT_SUBSTR)]

    if not hits:
        # try first significant English chunk
        eng = re.findall(r"[A-Za-z][A-Za-z0-9 /&.-]{2,}", name)
        for chunk in eng[:3]:
            hits.extend(rg_search(re.escape(chunk.strip()[:30]), limit=3))
        hits = [h for h in dict.fromkeys(hits) if not any(ig in h for ig in IGNORE_HIT_SUBSTR)]

    if not hits:
        return "غير موجود إطلاقًا"

    impl = [h for h in hits if any(w in h for w in WORKING_FILES)]
    evidence = (impl[0] if impl else hits[0]).replace(str(ROOT) + "/", "")

    if any(k in nl for k in PARTIAL_KW):
        return f"مبني جزئيًا — {evidence}"

    # broad umbrella concepts
    broad = ("data engine", "risk intelligence", "market intelligence", "decision intelligence", "trust os")
    if any(b in nl for b in broad):
        subs = ", ".join(h.replace(str(ROOT) + "/", "") for h in impl[:4]) or evidence
        return f"مبني جزئيًا — مفهوم عام؛ أسطح فرعية: {subs}"

    if "test" in " ".join(hits).lower() or any("/tests/" in h for h in hits):
        return f"مبني وشغال فعليًا — {evidence}"

    if len(impl) >= 1:
        return f"مبني جزئيًا — {evidence}"

    return f"غير مؤكَّد — إشارات فقط في {evidence}"


def discover_extra(existing_names: set[str]) -> list[dict[str, Any]]:
    """Features in repo with code evidence but absent from PDF names."""
    extras: list[dict[str, Any]] = []
    candidates = [
        ("Net-Edge Truth Layer", "net_edge_truth.py", "مبني وشغال فعليًا"),
        ("Proof Arena / Anti-Hype Mode", "proof_arena.py", "مبني وشغال فعليًا"),
        ("Decision Certificate DD Export", "decision_certificate.py", "مبني وشغال فعليًا"),
        ("Intelligence Ledger (execution intel)", "bd_platform/intelligence_ledger.py", "مبني وشغال فعليًا"),
        ("Squeeze Trigger Engine", "bd_platform/squeeze_trigger_engine.py", "مبني جزئيًا"),
        ("CEX-DEX Arbitrage Scanner", "bd_platform/cex_dex_arbitrage.py", "مبني وشغال فعليًا"),
        ("Coverage Honesty Board", "coverage_honesty.py", "مبني وشغال فعليًا"),
        ("Cap646/978 Institutional Closure Gates", "cap646/institutional_gate.py", "مبني وشغال فعليًا"),
        ("Wave-01 Data Engine (blackdark/data)", "blackdark/data/", "مبني جزئيًا"),
        ("B2B WebSocket Hub", "b2b_websocket_hub.py", "مبني وشغال فعليًا"),
        ("Didit KYC Webhook Integration", "didit_kyc.py", "مبني جزئيًا"),
        ("FinBERT Sentiment", "bd_platform/finbert_sentiment.py", "مبني جزئيًا"),
        ("Regulatory Compliance Guard (public oracle sanitization)", "regulatory_compliance_guard.py", "مبني وشغال فعليًا"),
        ("Constitution Gates / Product Constitution", "constitution_gates.py", "مبني وشغال فعليًا"),
        ("LP IL Live Simulator", "lp_il_simulator.py", "مبني جزئيًا"),
        ("Glass Box Challenge", "glass_box_challenge.py", "مبني جزئيًا"),
    ]
    norm_existing = {re.sub(r"\s+", " ", n.lower()) for n in existing_names}
    next_id = 817
    for title, path, status in candidates:
        key = title.lower()
        if not any(key[:12] in n or n[:12] in key for n in norm_existing):
            p = ROOT / path.split("+")[0].strip().split()[0]
            if p.exists() or (ROOT / path.split()[0]).exists():
                extras.append(
                    {
                        "id": next_id,
                        "name": title,
                        "type": "غير مذكور في الملف المرفق",
                        "status": f"{status} — {path}",
                    }
                )
                next_id += 1
    return extras


def write_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "capabilities"
    headers = ["#", "اسم الميزة/القدرة", "النوع", "الحالة"]
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["id"])
        ws.cell(row=i, column=2, value=r["name"])
        ws.cell(row=i, column=3, value=r["type"])
        c = ws.cell(row=i, column=4, value=r["status"])
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 78
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    if not PDF.exists():
        raise SystemExit(f"Missing attached PDF: {PDF}")
    entries = parse_pdf()
    if len(entries) != 816:
        raise SystemExit(f"Expected 816 rows from PDF, got {len(entries)}")
    print(f"Parsed {len(entries)} capabilities from attached PDF")

    rows: list[dict[str, Any]] = []
    for e in entries:
        status = classify(e)
        rows.append({"id": e["id"], "name": e["name"], "type": e["type"], "status": status})

    extras = discover_extra({e["name"] for e in entries})
    rows.extend(extras)
    print(f"Added {len(extras)} repo-only capabilities not in PDF")

    write_xlsx(rows, OUT)

    from collections import Counter

    counts = Counter()
    for r in rows[:816]:
        s = r["status"]
        if s.startswith("مبني وشغال"):
            counts["working"] += 1
        elif s.startswith("مبني جزئي"):
            counts["partial"] += 1
        elif s.startswith("غير موجود"):
            counts["missing"] += 1
        else:
            counts["uncertain"] += 1
    print("Status counts (816 PDF items):", dict(counts))
    print(f"Wrote {OUT} ({len(rows)} total rows)")


if __name__ == "__main__":
    main()
