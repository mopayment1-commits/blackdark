#!/usr/bin/env python3
"""Match partial PDF rows with platform_api.py-only evidence to concrete routes."""

from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from pdf_capability_registry import execute_binding, execute_evidence
from scripts.upgrade_partial_capabilities import apply_to_xlsx, parse_row

XLSX = ROOT / "capabilities_checklist.xlsx"
PLATFORM_API = ROOT / "platform_api.py"


def _tokens(name: str) -> set[str]:
    name = re.sub(r"[^\w\s\u0600-\u06FF]", " ", name.lower())
    parts = [p for p in name.split() if len(p) >= 3]
    stop = {"the", "and", "for", "with", "api", "data", "engine", "intelligence", "platform", "free"}
  # noqa: E501
    return {p for p in parts if p not in stop}


async def _build_route_index() -> list[dict]:
    text = PLATFORM_API.read_text(encoding="utf-8")
    routes: list[dict] = []
    blocks = re.split(r"(?=@router\.(get|post|put|delete)\()", text)
    i = 1
    while i < len(blocks):
        method = blocks[i]
        chunk = blocks[i + 1] if i + 1 < len(blocks) else ""
        m_path = re.search(r'["\']([^"\']+)["\']', chunk)
        m_fn = re.search(r"async def ([a-zA-Z_][a-zA-Z0-9_]*)", chunk)
        imports = re.findall(
            r"from\s+([\w.]+)\s+import\s+([a-zA-Z_][a-zA-Z0-9_,\s]+)",
            chunk,
        )
        func_calls = re.findall(r"return\s+([a-zA-Z_][a-zA-Z0-9_]*)_(\d+)\(", chunk)
        route = {
            "method": method,
            "path": m_path.group(1) if m_path else "",
            "handler": m_fn.group(1) if m_fn else "",
            "imports": imports,
            "layer_calls": func_calls,
            "text": chunk[:500].lower(),
        }
        routes.append(route)
        i += 2
    return routes


def _score_route(name: str, route: dict) -> int:
    score = 0
    toks = _tokens(name)
    hay = " ".join([route["path"], route["handler"], route["text"]]).lower()
    for tok in toks:
        if tok in hay:
            score += 3
        if tok in route["path"].lower():
            score += 2
    return score


async def _execute_route(route: dict) -> dict:
    if route["layer_calls"]:
        fn_base, cid = route["layer_calls"][0]
        from pdf_capability_registry import discover_bindings

        binding = discover_bindings().get(int(cid))
        if binding:
            return await execute_binding(binding[0], binding[1], capability_id=int(cid))
    for mod_path, names in route["imports"]:
        for name in [n.strip() for n in names.split(",")]:
            if name.startswith("_"):
                continue
            result = await execute_binding(mod_path, name)
            if result.get("ok"):
                return result
    return {"ok": False, "error": "route_exec_failed"}


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    routes = await _build_route_index()
    wb = load_workbook(XLSX)
    upgrades: dict[int, tuple[str, str]] = {}

    for row in wb.active.iter_rows(min_row=2, values_only=True):
        st, ev = parse_row(row[3])
        if st != "مبني جزئيًا":
            continue
        if "platform_api.py" not in ev:
            continue
        cid = int(row[0])
        name = str(row[1] or "")
        ranked = sorted((( _score_route(name, r), r) for r in routes), reverse=True, key=lambda x: x[0])
        if not ranked or ranked[0][0] < 1:
            continue
        best = ranked[0][1]
        result = await _execute_route(best)
        if result.get("ok"):
            binding = result.get("binding") or f"platform_api:{best['path']}"
            upgrades[cid] = ("مبني وشغال فعليًا", f"{binding} via platform_api{best['path']}")

    print(json.dumps({"upgraded": len(upgrades), "routes_indexed": len(routes)}, indent=2))
    if not args.dry_run and upgrades:
        apply_to_xlsx(upgrades)
        print(f"Updated {len(upgrades)} rows")


if __name__ == "__main__":
    asyncio.run(main())
