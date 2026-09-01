#!/usr/bin/env python3
"""Update capabilities_checklist.xlsx for official batch02 closure (IDs 51–100)."""

from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

XLSX = ROOT / "capabilities_checklist.xlsx"
RTM = ROOT / "docs" / "BATCH02_OFFICIAL_RTM_51_100.json"
STATUS_LABEL = "PRODUCTION-ALIGNED (official batch02)"


def main() -> None:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl required")

    rtm = json.loads(RTM.read_text(encoding="utf-8"))
    per_id = {int(k): v for k, v in rtm["per_id"].items()}

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    id_col = headers.index("#") + 1
    status_col = headers.index("الحالة") + 1

    updated = 0
    for row in ws.iter_rows(min_row=2):
        cid = row[id_col - 1].value
        if cid is None:
            continue
        cid = int(cid)
        if cid not in per_id:
            continue
        if per_id[cid]["status"] == "PRODUCTION-ALIGNED":
            row[status_col - 1].value = STATUS_LABEL
            updated += 1

    wb.save(XLSX)
    print(json.dumps({"updated_rows": updated, "path": str(XLSX)}, indent=2))


if __name__ == "__main__":
    main()
