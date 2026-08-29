#!/usr/bin/env python3
"""Build capabilities_checklist.xlsx — strict code audit for CAP978 IDs 1–816."""

from __future__ import annotations

import asyncio
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bd_platform.free_tier_capabilities import FREE_TIER_BASE_IDS, FREE_TIER_EXTENSION_IDS
from cap646.backend_registry import binding_for
from cap646.catalog import matrix_by_id
from cap646.runtime import VERIFIED_IDS, execute_capability
from cap646.waves import EXTERNAL_EVIDENCE_SLOTS, SIGNED_INFRA_SLOTS
from cap978.catalog import canonical_id, catalog_by_id, is_duplicate
from cap978.verify import execute_extension, verify_functional_978

OUT = ROOT / "capabilities_checklist.xlsx"
CATALOG = ROOT / "docs" / "cap978" / "CAP978_CATALOG.json"

# Handler fallbacks — not dedicated feature implementations
CATCH_ALL_SURFACES = frozenset(
    {
        "onchain_intelligence",
        "ai_decision_intelligence",
        "data_platform_storage",
        "market_data",
        "execution_trading",
        "institutional_ops",
        "alerts_workflow",
        "platform_codepath",
        "generic",
        "unknown",
    }
)

BROAD_CONCEPTS = (
    "data engine",
    "risk intelligence",
    "decision intelligence engine",
    "market intelligence",
    "platform intelligence",
)


def _loc(module: str | None, entry: str | None) -> str:
    if module and entry:
        return f"{module}.{entry}"
    return module or ""


async def _execute_row(capability_id: int) -> dict:
    params = {"symbol": "BTC", "tier": "whale", "coin_id": "bitcoin", "address": "0x0000000000000000000000000000000000000001"}
    user = {"email": "audit@blackdark.local", "tier": "elite"}
    try:
        if capability_id <= 646:
            if is_duplicate(capability_id):
                cid = canonical_id(capability_id)
                return await execute_capability(cid, skip_entitlement=True, params=params)
            return await execute_capability(capability_id, skip_entitlement=True, user=user, params=params)
        return await execute_extension(capability_id, user=user, params=params)
    except Exception as exc:
        return {"success": False, "error": str(exc)}


def _strict_status(
    capability_id: int,
    *,
    verify: dict,
    result: dict,
    gap: dict,
    binding: dict,
) -> str:
    verdict = verify.get("verdict", "UNKNOWN")
    checks = verify.get("checks") or {}
    failure = verify.get("failure_reason")
    surface = str(result.get("surface") or "")
    binding_source = str(result.get("binding_source") or binding.get("binding_source") or "")
    backend = result.get("backend_module") or binding.get("backend_module")
    entry = result.get("backend_entrypoint") or binding.get("backend_entrypoint")
    loc = _loc(backend, entry)
    binding_loc = _loc(binding.get("backend_module"), binding.get("backend_entrypoint"))
    components = gap.get("existing_code_components") or []
    tests = gap.get("tests_evidence") or []
    cap_name = gap.get("capability") or catalog_by_id().get(capability_id, {}).get("capability", "")
    gap_class = gap.get("final_classification", "")

    if verdict == "CANONICALLY_COVERED":
        canon = canonical_id(capability_id)
        return f"مبني وشغال فعليًا — مكرر/مغطى عبر القدرة الأساسية #{canon}"

    if verdict == "EXTERNAL_BLOCKED":
        if binding_loc:
            return f"مبني جزئيًا — محجوب خارجياً (vendor/rights)؛ ربط مسجل: {binding_loc}"
        return "مبني جزئيًا — محجوب خارجياً بدون تنفيذ داخلي مؤكد"

    if verdict == "EXTERNAL_EVIDENCE_REQUIRED" or capability_id in EXTERNAL_EVIDENCE_SLOTS | SIGNED_INFRA_SLOTS:
        miss = "دليل/شهادة خارجية مطلوبة"
        return f"مبني جزئيًا — {miss}" + (f"؛ كود: {loc or binding_loc}" if (loc or binding_loc) else "")

    if verdict == "FUNCTIONALLY_INCOMPLETE":
        miss = failure or "؛ ".join(k for k, v in checks.items() if v is False) or "فحص وظيفي غير مكتمل"
        return f"مبني جزئيًا — {miss}" + (f"؛ كود: {loc or binding_loc}" if (loc or binding_loc) else "")

    # Free-tier surfaces execute outside generic handler fallbacks
    if capability_id in FREE_TIER_BASE_IDS | FREE_TIER_EXTENSION_IDS:
        code_loc = loc or binding_loc or "bd_platform.free_tier_capabilities"
        if bool(result.get("success")):
            return f"مبني وشغال فعليًا — {code_loc}"
        return f"مبني جزئيًا — كود موجود ({code_loc}) لكن التحقق الحي فشل (API/شبكة غير متاحة في بيئة الفحص)"

    if capability_id in VERIFIED_IDS and surface and surface not in CATCH_ALL_SURFACES:
        return f"مبني وشغال فعليًا — {loc or binding_loc} ({surface})"

    # Generic handler fallback — partial even if verify says complete
    if surface in CATCH_ALL_SURFACES:
        hint = loc or binding_loc or surface
        extra = ""
        if binding_loc and binding_loc != loc:
            extra = f"؛ ربط مسجل (غير مفعّل في المسار الحي): {binding_loc}"
        if components:
            extra += f"؛ مكونات جزئية: {', '.join(components[:2])}"
        return f"مبني جزئيًا — توجيه عام ({surface}) بدون منطق مخصص للقدرة؛ المسار الحي: {hint}{extra}"

    if binding_source == "track_default" and not components:
        return f"مبني جزئيًا — ربط افتراضي للمسار فقط ({binding_loc or loc})"

    if bool(result.get("success")) and (loc or binding_loc):
        extra = ""
        if components:
            extra = f"؛ مكونات: {', '.join(components[:2])}"
        if tests:
            extra += f"؛ اختبارات: {', '.join(tests[:2])}"
        if any(term in cap_name.lower() for term in BROAD_CONCEPTS):
            return f"مبني جزئيًا — مفهوم عام؛ سطح فرعي موجود: {loc or binding_loc} ({surface}){extra}"
        return f"مبني وشغال فعليًا — {loc or binding_loc} ({surface}){extra}"

    if components:
        return f"مبني جزئيًا — مكونات موجودة لكن التنفيذ غير مؤكد: {', '.join(components[:2])}"

    if gap_class == "NOT_IMPLEMENTED" and not bool(result.get("success")):
        return "غير موجود إطلاقًا"

    if gap_class == "NOT_IMPLEMENTED":
        return "غير موجود إطلاقًا"

    return "غير مؤكَّد"


async def audit_one(capability_id: int) -> dict:
    row = catalog_by_id().get(capability_id, {})
    gap = matrix_by_id().get(capability_id, {}) if capability_id <= 646 else {}
    verify = await verify_functional_978(
        capability_id,
        user={"email": "audit@blackdark.local", "tier": "elite"},
    )
    result = await _execute_row(capability_id)
    try:
        if capability_id <= 646:
            binding = binding_for(capability_id)
        else:
            from cap978.extension_registry import resolve_extension_binding

            eb = resolve_extension_binding(capability_id)
            binding = {
                "backend_module": eb.module,
                "backend_entrypoint": eb.entrypoint,
                "surface": eb.surface,
                "binding_source": eb.source,
            }
    except Exception:
        binding = {
            "backend_module": result.get("backend_module"),
            "backend_entrypoint": result.get("backend_entrypoint"),
            "surface": result.get("surface"),
            "binding_source": result.get("binding_source"),
        }

    status = _strict_status(capability_id, verify=verify, result=result, gap=gap, binding=binding)
    return {
        "id": capability_id,
        "track": row.get("track", ""),
        "track_name": row.get("track_name", ""),
        "capability": row.get("capability", ""),
        "scope": row.get("scope", ""),
        "الحالة": status,
    }


async def run_audit(ids: list[int]) -> list[dict]:
    sem = asyncio.Semaphore(6)
    rows: list[dict] = []

    async def _one(cid: int) -> dict:
        async with sem:
            return await audit_one(cid)

    tasks = [_one(cid) for cid in ids]
    for i, coro in enumerate(asyncio.as_completed(tasks)):
        rows.append(await coro)
        if (i + 1) % 100 == 0:
            print(f"audited {i + 1}/{len(ids)}", flush=True)
    rows.sort(key=lambda r: r["id"])
    return rows


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "capabilities"
    headers = ["#", "المسار", "اسم المسار", "القدرة", "النطاق", "الحالة"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["id"])
        ws.cell(row=i, column=2, value=r["track"])
        ws.cell(row=i, column=3, value=r["track_name"])
        ws.cell(row=i, column=4, value=r["capability"])
        ws.cell(row=i, column=5, value=r["scope"])
        c = ws.cell(row=i, column=6, value=r["الحالة"])
        c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 78
    ws.freeze_panes = "A2"
    wb.save(path)


async def main() -> None:
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    ids = [int(r["id"]) for r in catalog if int(r["id"]) <= 816]
    assert len(ids) == 816
    print(f"Auditing {len(ids)} capabilities (strict mode) …")
    rows = await run_audit(ids)
    write_xlsx(rows, OUT)
    from collections import Counter

    counts = Counter()
    for r in rows:
        s = r["الحالة"]
        if s.startswith("مبني وشغال"):
            counts["مبني وشغال فعليًا"] += 1
        elif s.startswith("مبني جزئيًا"):
            counts["مبني جزئيًا"] += 1
        elif s.startswith("غير موجود"):
            counts["غير موجود إطلاقًا"] += 1
        else:
            counts["غير مؤكَّد"] += 1
    print(json.dumps(dict(counts), ensure_ascii=False, indent=2))
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    asyncio.run(main())
