#!/usr/bin/env python3
"""Upgrade partial PDF capabilities when dedicated binding executes successfully."""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from pdf_capability_registry import batch_test_module_for, discover_bindings, execute_capability, execute_evidence, resolve_evidence_binding

XLSX = ROOT / "capabilities_checklist.xlsx"
OUT = ROOT / "capabilities_checklist.xlsx"


def parse_row(cell: str | None) -> tuple[str, str]:
    if not cell:
        return "غير مؤكد", ""
    s = str(cell)
    for prefix in ("مبني وشغال فعليًا", "مبني جزئيًا", "غير موجود إطلاقًا", "غير مؤكد"):
        if s.startswith(prefix):
            return prefix, s[len(prefix) :].lstrip(" —-")
    return "غير مؤكد", s


def format_status(status: str, evidence: str) -> str:
    return f"{status} — {evidence}" if evidence else status


async def upgrade_batch(ids: list[int], *, evidence_map: dict[int, str] | None = None, dry_run: bool = False) -> list[dict]:
    results = []
    evidence_map = evidence_map or {}
    for cid in ids:
        evidence = evidence_map.get(cid, "")
        exec_result = await execute_evidence(cid, evidence)
        binding = discover_bindings().get(cid) or resolve_evidence_binding(evidence)
        test_mod = batch_test_module_for(cid)
        ok = bool(exec_result.get("ok"))
        if ok and binding:
            evidence_out = f"{binding[0]}.{binding[1]}"
            if test_mod:
                evidence_out += f" + {test_mod}"
            results.append(
                {
                    "id": cid,
                    "status": "مبني وشغال فعليًا",
                    "evidence": evidence_out,
                    "exec": exec_result,
                }
            )
        else:
            results.append({"id": cid, "status": "مبني جزئيًا", "skipped": True, "exec": exec_result})
    return results


def apply_to_xlsx(updates: dict[int, tuple[str, str]]) -> None:
    wb = load_workbook(XLSX)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        cid = row[0].value
        if cid is None:
            continue
        cid = int(cid)
        if cid in updates:
            st, ev = updates[cid]
            row[3].value = format_status(st, ev)
            row[3].alignment = Alignment(wrap_text=True)
    wb.save(XLSX)


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0, help="Max IDs to process (0=all partial)")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--ids", type=str, default="", help="Comma-separated IDs")
    args = parser.parse_args()

    wb = load_workbook(XLSX)
    partial_ids: list[int] = []
    evidence_map: dict[int, str] = {}
    for r in wb.active.iter_rows(min_row=2, values_only=True):
        st, ev = parse_row(r[3])
        if st == "مبني جزئيًا":
            cid = int(r[0])
            partial_ids.append(cid)
            evidence_map[cid] = ev

    if args.ids:
        partial_ids = [int(x.strip()) for x in args.ids.split(",") if x.strip()]

    # prioritize IDs with discovered bindings
    bindings = discover_bindings()
    partial_ids.sort(key=lambda i: (0 if i in bindings else 1, i))
    if args.limit:
        partial_ids = partial_ids[: args.limit]

    print(f"Processing {len(partial_ids)} partial capabilities …", flush=True)
    upgrades: dict[int, tuple[str, str]] = {}
    ok_count = 0
    for cid in partial_ids:
        res = (await upgrade_batch([cid], evidence_map=evidence_map))[0]
        if res.get("status") == "مبني وشغال فعليًا":
            upgrades[cid] = (res["status"], res["evidence"])
            ok_count += 1
        if ok_count and ok_count % 50 == 0:
            print(f"  upgraded {ok_count} …", flush=True)

    print(json.dumps({"upgraded": ok_count, "processed": len(partial_ids)}, indent=2))
    if not args.dry_run and upgrades:
        apply_to_xlsx(upgrades)
        print(f"Updated {len(upgrades)} rows in {XLSX}")


if __name__ == "__main__":
    asyncio.run(main())
