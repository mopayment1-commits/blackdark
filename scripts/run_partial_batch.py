#!/usr/bin/env python3
"""Run a partial-capability batch manifest (verify + optional xlsx upgrade)."""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pdf_capability_registry import batch_test_module_for, discover_bindings, execute_capability
from scripts.upgrade_partial_capabilities import apply_to_xlsx, parse_row, upgrade_batch

BATCH_DIR = Path(__file__).resolve().parent / "partial_batches"


def load_batch(name: str) -> dict:
    path = BATCH_DIR / f"{name}.json"
    if not path.is_file():
        raise FileNotFoundError(path)
    return json.loads(path.read_text(encoding="utf-8"))


async def run_batch(name: str, *, dry_run: bool = False, upgrade: bool = False) -> dict:
    manifest = load_batch(name)
    ids = [int(x) for x in manifest["capability_ids"]]
    bindings = discover_bindings()

    results: list[dict] = []
    ok = fail = skip = 0
    for cid in ids:
        if cid not in bindings:
            skip += 1
            results.append({"id": cid, "status": "skip_no_binding"})
            continue
        exec_result = await execute_capability(cid)
        passed = bool(exec_result.get("ok"))
        if passed:
            ok += 1
            results.append({"id": cid, "status": "ok", "binding": f"{bindings[cid][0]}.{bindings[cid][1]}"})
        else:
            fail += 1
            results.append({"id": cid, "status": "fail", "error": exec_result.get("error"), "exec": exec_result})

    summary = {
        "batch": manifest.get("batch"),
        "label": manifest.get("label"),
        "processed": len(ids),
        "ok": ok,
        "fail": fail,
        "skip_no_binding": skip,
        "with_batch_test": sum(1 for cid in ids if batch_test_module_for(cid)),
    }

    if upgrade and not dry_run:
        from openpyxl import load_workbook

        wb = load_workbook(ROOT / "capabilities_checklist.xlsx")
        evidence_map: dict[int, str] = {}
        partial_ids: list[int] = []
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            st, ev = parse_row(row[3])
            cid = int(row[0])
            if st == "مبني جزئيًا" and cid in ids:
                partial_ids.append(cid)
                evidence_map[cid] = ev
        upgrades: dict[int, tuple[str, str]] = {}
        for cid in partial_ids:
            res = (await upgrade_batch([cid], evidence_map=evidence_map))[0]
            if res.get("status") == "مبني وشغال فعليًا":
                upgrades[cid] = (res["status"], res["evidence"])
        if upgrades:
            apply_to_xlsx(upgrades)
        summary["xlsx_upgraded"] = len(upgrades)

    summary["results_sample"] = [r for r in results if r["status"] != "ok"][:15]
    return summary


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("batch", help="Manifest name without .json (e.g. batch_01)")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--upgrade", action="store_true", help="Promote still-partial xlsx rows when exec ok")
    args = parser.parse_args()
    summary = await run_batch(args.batch, dry_run=args.dry_run, upgrade=args.upgrade)
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    asyncio.run(main())
