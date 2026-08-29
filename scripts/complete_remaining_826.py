#!/usr/bin/env python3
"""Complete remaining PDF capabilities — upgrade all except human-only external slots."""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from pdf_capability_registry import batch_test_module_for, discover_bindings, execute_capability
from scripts.upgrade_partial_capabilities import apply_to_xlsx, format_status, parse_row

XLSX = ROOT / "capabilities_checklist.xlsx"

# External vendor / human provisioning only — remain مبني جزئيًا with honest label
HUMAN_ONLY_IDS: frozenset[int] = frozenset(
    {
        693,  # Polygon.io paid API license — requires external contract
    }
)

HUMAN_ONLY_EVIDENCE: dict[int, str] = {
    693: "EXTERNAL_BLOCKED — Polygon.io API license requires human vendor contract; proxy: bd_platform/oneinch_connector.py",
}


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(XLSX)
    partial_ids: list[int] = []
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        st, _ = parse_row(row[3])
        if st == "مبني جزئيًا":
            partial_ids.append(int(row[0]))

    upgrades: dict[int, tuple[str, str]] = {}
    human_labeled: dict[int, tuple[str, str]] = {}

    for cid in partial_ids:
        if cid in HUMAN_ONLY_IDS:
            human_labeled[cid] = ("مبني جزئيًا", HUMAN_ONLY_EVIDENCE.get(cid, "EXTERNAL_BLOCKED — human provisioning required"))
            continue
        result = await execute_capability(cid)
        binding = discover_bindings().get(cid)
        if result.get("ok") and binding:
            evidence = f"{binding[0]}.{binding[1]}"
            test_mod = batch_test_module_for(cid)
            if test_mod:
                evidence += f" + {test_mod}"
            upgrades[cid] = ("مبني وشغال فعليًا", evidence)

    print(
        json.dumps(
            {
                "upgraded": len(upgrades),
                "human_only_labeled": len(human_labeled),
                "still_partial": len(partial_ids) - len(upgrades) - len(human_labeled),
                "processed": len(partial_ids),
            },
            indent=2,
        )
    )

    if args.dry_run:
        return

    merged = {**upgrades, **human_labeled}
    if merged:
        apply_to_xlsx(merged)
        print(f"Updated {len(merged)} rows in {XLSX}")


if __name__ == "__main__":
    asyncio.run(main())
